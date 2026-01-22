import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import json
from sqlalchemy import create_engine, text
import getpass
from pandastable import Table
import smtplib
from email.message import EmailMessage
import sys
from pathlib import Path
import os
import ast
from PIL import Image, ImageTk
# import yaml  # 暫時停用 YAML 功能
import openpyxl
from datetime import datetime
import re

# 添加 lib 目錄到路徑
sys.path.insert(0, str(Path(__file__).parent.parent.parent / 'lib'))
import mainlib

# 1. 程式開始時必須的公用函式與設定
# --------------------------------------------------
# messagebox.showinfo("歡迎", "歡迎使用建大ERP程式！\n\n程式已啟動，除錯模式已開啟。")
# print("=== ERP程式啟動成功！===")

# print("如果您點擊查詢按鈕，將會看到詳細的除錯訊息。")
# 獲取 AD 登入使用者名稱
username = getpass.getuser()

# 使用 lib 提供的路徑管理函數自動定位到專案根目錄的 xlsx 資料夾
subno_path = mainlib.get_xlsx_path("subno.xlsx")
input_path = mainlib.get_xlsx_path("input.xlsx")

# 載入 subno 和 SQL 模板資料
try:
    df_factories = pd.read_excel(subno_path, sheet_name="subno", engine="openpyxl")
    factories = df_factories["subno"].tolist()
    connections = dict(zip(df_factories["subno"], df_factories["connection"]))

    df_sql = pd.read_excel(input_path, sheet_name="sql", engine="openpyxl")
    
    # 修正欄位名稱（如果第一欄不是 'code'，重新命名）
    if df_sql.columns[0] != 'code':
        df_sql.rename(columns={df_sql.columns[0]: 'code'}, inplace=True)
    
    # 硬編碼加入 D002 定義（支援單廠與集團統計）
    if 'D002' not in df_sql['code'].values:
        d002_row = pd.DataFrame([{
            'code': 'D002',
            'Chinese': "['年月起(YYYYMM)', '年月迄(YYYYMM)']",
            'input': "['input1', 'input2']",
            'input_width': '[6, 6]',
            'sql': 'D002_SPECIAL',
            'note': '單廠/集團統計-按產品類型'
        }])
        df_sql = pd.concat([df_sql, d002_row], ignore_index=True)
    
    sql_templates = dict(zip(df_sql["code"], df_sql["sql"]))
    input_labels = dict(zip(df_sql["code"], df_sql["Chinese"]))
    input_keys = dict(zip(df_sql["code"], df_sql["input"]))
    input_widths = dict(zip(df_sql["code"], df_sql["input_width"]))
    notes = dict(zip(df_sql["code"], df_sql["note"]))
    
except FileNotFoundError as e:
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("檔案遺失", f"找不到必要的檔案: {e.filename}")
    sys.exit()

# GUI 設定
root = tk.Tk()
root.title("ERP SQL 查詢工具")
root.state('zoomed')

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# 變數宣告
df_result = None
input_entries = {}
code_display = [f"{row['code']} - {row['Chinese']}" for _, row in df_sql.iterrows()]
code_map = dict(zip(code_display, df_sql["code"]))
subno_var = tk.StringVar(value="T.天津廠")
code_var = tk.StringVar()

# 2. factory_frame: 廠別與查詢代碼的元件與函式
# --------------------------------------------------

def convert_date_to_epoch(date_str):
    """
    將 YYYYMMDD 格式的日期轉換為 epoch timestamp (秒)
    例如：'20240101' -> 1704038400 (2024-01-01 00:00:00的epoch時間)
    """
    try:
        # 檢查是否為8位數字格式
        if len(date_str) == 8 and date_str.isdigit():
            # 解析日期
            dt = datetime.strptime(date_str, '%Y%m%d')
            # 轉換為 epoch timestamp (秒)
            epoch = int(dt.timestamp())
            return epoch
        return None
    except Exception as e:
        print(f"日期轉換錯誤: {e}")
        return None

def is_date_parameter(param_name):
    """
    判斷參數名稱是否為日期參數
    可根據實際需求擴充判斷邏輯
    """
    # 常見的日期參數名稱
    date_keywords = ['date', 'input1', 'input2', 'start', 'end', 'begin', 'finish']
    param_lower = param_name.lower()
    return any(keyword in param_lower for keyword in date_keywords)

def test_connection(event=None):
    """測試選擇的廠別資料庫連線是否正常。"""
    subno = subno_var.get()
    
    if subno not in connections:
        message_label.config(text=f"廠別 '{subno}' 不在連線清單中", fg="red")
        return
    
    try:
        conn_info = json.loads(connections[subno])
        host = conn_info['host']
        dbname = conn_info['dbname']
        user = conn_info['user']
        
        # 嘗試建立資料庫連線
        conn_str = f"postgresql://{conn_info['user']}:{conn_info['password']}@{host}/{dbname}"
        engine = create_engine(conn_str)
        
        # 測試連線 - 執行簡單查詢
        with engine.connect() as connection:
            connection.execute(text("SELECT 1"))
        
        # 連線成功
        message_label.config(
            text=f"✓ 連線正常 | Host: {host} | DB: {dbname} | User: {user}",
            fg="green"
        )
        
    except Exception as e:
        # 連線失敗
        try:
            conn_info = json.loads(connections[subno])
            host = conn_info['host']
            dbname = conn_info['dbname']
            user = conn_info['user']
            message_label.config(
                text=f"✗ 無法連線 | Host: {host} | DB: {dbname} | User: {user}",
                fg="red"
            )
        except:
            message_label.config(text=f"連線測試失敗: {e}", fg="red")

def update_title(event=None):
    """根據選擇的查詢代碼，更新主視窗標題。"""
    selected_display = code_var.get()
    selected_code = code_map.get(selected_display)
    if selected_code:
        title_text = f"{selected_code} - {input_labels.get(selected_code, '')}"
        root.title(title_text)

factory tk.Frame(root, borderwidth=1, relief="solid")
factory_frame.place(x=0, y=0, width=screen_width, height=40)

tk.Label(factory_frame, text="廠別", font=("Arial", 14)).place(x=5, y=5)
subno_combo = ttk.Combobox(factory_frame, textvariable=subno_var, values=factories, state="readonly", font=("Arial", 14))
subno_combo.place(x=80, y=5)

tk.Label(factory_frame, text="查詢代碼", font=("Arial", 14)).place(x=250, y=5)
code_combo = ttk.Combobox(factory_frame, textvariable=code_var, values=code_display, state="readonly", width=80, font=("Arial", 14))
code_combo.place(x=380, y=5)

subno_combo.bind("<<ComboboxSelected>>", test_connection)
code_combo.bind("<<ComboboxSelected>>", update_title)
update_title()
if code_display:
    code_var.set(code_display[0])

# 3. error_frame: 錯誤與訊息顯示的元件與函式
# --------------------------------------------------

def show_sql():
    """在訊息標籤中顯示 SQL 語法。"""
    code_key = code_map.get(code_var.get())
    if code_key in sql_templates and pd.notna(sql_templates[code_key]):
        message_label.config(text=f"SQL語法：\n{sql_templates[code_key]}", fg="blue")
    else:
        message_label.config(text="SQL語法：無", fg="red")

def show_note():
    """在訊息標籤中顯示查詢條件說明。"""
    code_key = code_map.get(code_var.get())
    if code_key in notes and pd.notna(notes[code_key]):
        message_label.config(text=f"條件說明：\n{notes[code_key]}", fg="red")
    else:
        message_label.config(text="條件說明：無", fg="red")

error_frame = tk.Frame(root, borderwidth=2, relief="solid", bg="lightgray")
error_frame.pack(side="bottom", fill="x", padx=0, pady=0)

message_label = tk.Label(error_frame, text="系統訊息區域", font=("Arial", 14), fg="red", anchor="w", justify="left", bg="lightgray", height=3)
message_label.pack(side="left", fill="both", expand=True, padx=5, pady=5)

# 進度顯示標籤（靠右）
progress_label = tk.Label(error_frame, text="", font=("Courier New", 14, "bold"), fg="blue", anchor="e", bg="lightgray", width=25, height=3)
progress_label.pack(side="right", padx=10, pady=5)

# 進度更新函數
def update_progress(percent):
    """更新進度條，使用方塊圖顯示"""
    filled_blocks = percent // 5  # 1個方塊 = 5%
    empty_blocks = 20 - filled_blocks
    progress_bar = "█" * filled_blocks + "░" * empty_blocks
    if percent < 100:
        progress_label.config(text=f"執行中 {percent}% {progress_bar}", fg="blue")
    else:
        progress_label.config(text=f"完成 {percent}% {progress_bar}", fg="green")
    root.update()
    root.after(50)  # 延遲50ms，確保GUI更新

# 4. input_frame: 輸入欄位的元件與函式
# --------------------------------------------------

def update_inputs(event=None):
    """根據選擇的查詢代碼，動態更新輸入欄位並顯示圖片。"""
    for widget in input_frame.winfo_children():
        widget.destroy()
    input_entries.clear()
    code_key = code_map.get(code_var.get())

    if code_key in input_labels and code_key in input_keys:
        try:
            labels = ast.literal_eval(input_labels[code_key])
            keys = ast.literal_eval(input_keys[code_key])
            widths = ast.literal_eval(input_widths.get(code_key, '[]'))
            if len(widths) != len(labels):
                widths = [20] * len(labels)
            
            current_x = 5
            for i, (label_text, key) in enumerate(zip(labels, keys)):
                label = tk.Label(input_frame, text=label_text, font=("Arial", 14))
                label.place(x=current_x, y=5)
                label_width = label.winfo_reqwidth()
                current_x += label_width + 5
                
                entry_width = widths[i] if i < len(widths) else 20
                entry = tk.Entry(input_frame, font=("Arial", 14), width=entry_width)
                entry.place(x=current_x, y=5)
                input_entries[key] = entry
                
                current_x += entry.winfo_reqwidth() + 20
                
        except Exception as e:
            message_label.config(text=f"解析輸入欄位失敗: {e}", fg="red")

    if code_key in notes and pd.notna(notes[code_key]):
        message_label.config(text=f"查詢說明：{notes[code_key]}", fg="blue")
    else:
        message_label.config(text="查詢說明：無", fg="blue")

    # 新增圖片顯示功能
    for widget in result_frame.winfo_children():
        widget.destroy()

    subno_prefix = subno_var.get().split('.')[0]
    image_name = f"{subno_prefix}-{code_key}.png"
    png_path = mainlib.get_image_path(image_name)
    
    if png_path.exists():
        try:
            img = Image.open(png_path)
            # 調整圖片大小以適應 result_frame
            frame_width = result_frame.winfo_width()
            frame_height = result_frame.winfo_height()
            img.thumbnail((frame_width, frame_height), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            
            img_label = tk.Label(result_frame, image=photo)
            img_label.image = photo  # 保持引用以防止垃圾回收
            img_label.pack(expand=True, fill="both")
            
        except Exception as e:
            message_label.config(text=f"圖片載入失敗: {e}", fg="red")
    else:
        message_label.config(text=f"找不到圖片檔案: {image_name}", fg="red")
        
input_frame = tk.Frame(root, borderwidth=1, relief="solid")
input_frame.place(x=0, y=45, width=screen_width, height=40)

# 5. result_frame: 查詢結果顯示的元件與函式
# --------------------------------------------------

def format_numeric_columns(df):
    """
    格式化 DataFrame 中的數字欄位：
    - cnt: 整數格式，靠右對齊
    - quantity_sum: 固定2位小數，靠右對齊
    - 其他欄位：如果是整數顯示整數，如果有小數根據最大小數位數格式化
    """
    if df is None or len(df) == 0:
        return df, {}
    
    df_formatted = df.copy()
    numeric_cols_info = {}  # 記錄數字欄位的格式資訊
    
    for col in df_formatted.columns:
        # 檢查是否為數字類型
        if pd.api.types.is_numeric_dtype(df_formatted[col]):
            # 排除 NaN 值
            non_null_values = df_formatted[col].dropna()
            
            if len(non_null_values) == 0:
                continue
            
            # 特殊處理 quantity_sum 欄位：固定2位小數
            if col == 'quantity_sum':
                # 先格式化所有數字為2位小數（不含千分位）
                formatted_values = [f"{float(x):.2f}" for x in non_null_values]
                max_length = max(len(v) for v in formatted_values)
                
                df_formatted[col] = df_formatted[col].apply(
                    lambda x: f"{float(x):.2f}".rjust(max_length) if pd.notna(x) else ''
                )
                numeric_cols_info[col] = 'decimal_2'
                continue
            
            # 檢查是否所有值都是整數（沒有小數部分）
            is_all_integer = all(non_null_values.apply(lambda x: float(x).is_integer()))
            
            if is_all_integer:
                # 全部是整數，轉換為字串格式（保持整數顯示）
                formatted_values = [f"{int(x)}" for x in non_null_values]
                max_length = max(len(v) for v in formatted_values)
                
                df_formatted[col] = df_formatted[col].apply(
                    lambda x: f"{int(x)}".rjust(max_length) if pd.notna(x) else ''
                )
                numeric_cols_info[col] = 'integer'
            else:
                # 有小數的情況，計算最大小數位數
                max_decimal_places = 0
                for value in non_null_values:
                    if pd.notna(value):
                        # 轉換為字串後分析小數位數
                        str_value = f"{float(value):.15f}".rstrip('0')  # 先保留足夠位數
                        if '.' in str_value:
                            decimal_part = str_value.split('.')[1]
                            decimal_places = len(decimal_part)
                            max_decimal_places = max(max_decimal_places, decimal_places)
                
                # 根據最大小數位數統一格式化（補零），轉為字串保持格式，並計算最大寬度實現右對齊
                if max_decimal_places > 0:
                    # 先格式化所有數字（不含千分位）
                    formatted_values = [f"{float(x):.{max_decimal_places}f}" for x in non_null_values]
                    max_length = max(len(v) for v in formatted_values)
                    
                    df_formatted[col] = df_formatted[col].apply(
                        lambda x: f"{float(x):.{max_decimal_places}f}".rjust(max_length) if pd.notna(x) else ''
                    )
                    numeric_cols_info[col] = f'decimal_{max_decimal_places}'
    
    return df_formatted, numeric_cols_info

def execute_sql():
    """執行 SQL 查詢並顯示結果。"""
    global df_result, is_group_query
    is_group_query = False  # 標記為單一廠別查詢

    # 基本檢查和除錯訊息
    print("\n" + "="*80)
    print("QUERY: 查詢按鈕被點擊！開始執行 SQL 查詢...")
    print("="*80)
    print(f"當前時間: {pd.Timestamp.now()}")
    try:
        update_progress(0)
        message_label.config(text="START: 開始執行查詢...", fg="red")
        root.update()
        print("SUCCESS: 訊息標籤已更新為：開始執行查詢...")

        # 強制GUI更新
        root.after(100, lambda: None)
    except Exception as e:
        print(f"ERROR: GUI更新錯誤: {e}")
    print("1111")
    
    update_progress(10)
    root.update()
    
    subno = subno_var.get()
    selected_code = code_var.get()
    code_key = code_map.get(selected_code)
    
    # 特殊處理：D002 單廠統計
    if code_key == 'D002':
        if not subno:
            messagebox.showerror("錯誤", "D002 需要選擇廠別！")
            return
        execute_d002(is_group=False)
        return

    message_label.config(text=f"廠別: {subno}, 代碼: {selected_code}, 鍵值: {code_key}", fg="blue")
    update_progress(20)
    root.update()

    

    if subno not in connections:
        message_label.config(text=f"廠別 '{subno}' 不在連線清單中", fg="red")
        progress_label.config(text="", fg="blue")
        return
    print('2222')
    
    update_progress(30)
    root.update()
    
    if code_key not in sql_templates:
        message_label.config(text=f"代碼鍵值 '{code_key}' 不在SQL模板中", fg="red")
        progress_label.config(text="", fg="blue")
        return

    # 檢查輸入欄位
    message_label.config(text=f"檢查輸入欄位，共 {len(input_entries)} 個欄位", fg="blue")
    update_progress(40)
    root.update()

    for key, entry in input_entries.items():
        value = entry.get()
        message_label.config(text=f"欄位 {key}: '{value}'", fg="blue")
        root.update()

    missing_inputs = [key for key, entry in input_entries.items() if not entry.get()]
    if missing_inputs:
        message_label.config(text=f"以下欄位尚未輸入: {', '.join(missing_inputs)}", fg="red")
        progress_label.config(text="", fg="blue")
        return

    try:
        update_progress(50)
        root.update()
        
        sql = sql_templates[code_key]
        message_label.config(text=f"原始SQL載入完成，廠別: {subno}", fg="blue")
        
        # ===== 終端機顯示：原始 SQL =====
        print("\n" + "=" * 100)
        print(f"【廠別】: {subno}")
        print(f"【查詢代碼】: {code_key}")
        print("=" * 100)
        print("【原始 SQL 語法】:")
        print(sql)
        print("=" * 100)
        
        update_progress(60)
        root.update()

        # 根據選擇的廠別替換 SQL 語法
        selected_factory = df_factories[df_factories["subno"] == subno]
        if selected_factory.empty:
            message_label.config(text="找不到對應的廠別設定", fg="red")
            print("❌ 錯誤：找不到對應的廠別設定")
            return
        replace1_val = selected_factory.iloc[0]["replace1"]
        replace2_val = selected_factory.iloc[0]["replace2"]

        # 替換 subno 和 factory 部分（陣列方式多次替換）
        if pd.notna(replace1_val):
            print(f"\n【替換 1】subno/factory (陣列方式):")
            print(f"  原始值: {replace1_val}")
            
            # 定義原始的替換陣列（支援大小寫）
            原始陣列 = [("'7' as subno", "'7' AS subno"), ("'L' as factory", "'L' AS factory")]
            
            # 將 replace1_val 拆分成陣列（用逗號分隔）
            try:
                # 移除首尾空白，然後用逗號分隔
                新陣列 = [item.strip() for item in replace1_val.split(',')]
                
                print(f"  拆分後的陣列: {新陣列}")
                
                # 確保陣列長度一致
                if len(新陣列) == len(原始陣列):
                    # 逐一替換（支援大小寫）
                    for i, (舊值組, 新值) in enumerate(zip(原始陣列, 新陣列), 1):
                        # 舊值組包含兩種格式：(小寫as, 大寫AS)
                        舊值_小寫, 舊值_大寫 = 舊值組
                        print(f"  替換 {i}: '{舊值_小寫}' 或 '{舊值_大寫}' → '{新值}'")
                        # 先替換小寫，再替換大寫
                        sql = sql.replace(舊值_小寫, 新值)
                        sql = sql.replace(舊值_大寫, 新值)
                    message_label.config(text=f"已替換 subno/factory (共{len(新陣列)}次)", fg="blue")
                else:
                    print(f"  ⚠ 警告：陣列長度不一致，使用原始方式替換")
                    sql = sql.replace("'7' as subno,'L' as factory", replace1_val)
                    
            except Exception as e:
                print(f"  ⚠ 警告：拆分失敗，使用原始方式替換 - {e}")
                sql = sql.replace("'7' as subno,'L' as factory", replace1_val)
                
            root.update()

        # 替換資料庫 schema 名稱
        if pd.notna(replace2_val):
            print(f"\n【替換 2】schema 名稱:")
            print(f"  舊值: kt2mes")
            print(f"  新值: {replace2_val}")
            sql = sql.replace("kt2mes", replace2_val)
            message_label.config(text=f"已替換 schema: kt2mes -> {replace2_val}", fg="blue")
            root.update()

        # 替換輸入參數 (優化日期處理)
        print(f"\n【替換 3】輸入參數:")
        date_params = {}  # 儲存日期參數用於後續優化
        
        for key, entry in input_entries.items():
            value = entry.get()
            if value:  # 只有在有值時才替換
                # 檢查是否為日期參數且為8位數字格式
                if is_date_parameter(key) and len(value) == 8 and value.isdigit():
                    epoch_value = convert_date_to_epoch(value)
                    if epoch_value:
                        date_params[key] = {'original': value, 'epoch': epoch_value}
                        print(f"  {key} -> '{value}' (epoch: {epoch_value})")
                    else:
                        print(f"  {key} -> '{value}'")
                else:
                    print(f"  {key} -> '{value}'")
                
                sql = sql.replace(key, f"'{value}'")
                message_label.config(text=f"已替換參數 {key}: {value}", fg="blue")
                root.update()
        
        # 優化 SQL：將日期查詢移到子查詢內部
        if date_params:
            print(f"\n【SQL 優化】檢查是否可以優化:")
            # 檢測並替換 BETWEEN 日期查詢
            if 'input1' in date_params and 'input2' in date_params:
                start_date = date_params['input1']['original']
                end_date = date_params['input2']['original']
                
                # 檢查是否為子查詢結構（包含 "FROM ( ... ) AS cr"）
                has_subquery = re.search(r'FROM\s*\(\s*SELECT.*?\)\s*AS\s+cr', sql, re.DOTALL | re.IGNORECASE)
                
                if has_subquery:
                    print(f"  ✓ 偵測到子查詢結構，可以優化")
                    
                    # 檢測 SQL 使用哪種時間欄位
                    uses_filter_time = 'cr.filter_time' in sql or 'filter_time AS' in sql
                    uses_formatted_time = 'cr.formatted_time' in sql
                    
                    if uses_filter_time:
                        print(f"  ✓ 偵測到使用 filter_time (YYYYMM 格式)")
                        time_field = 'filter_time'
                        time_format = 'YYYYMM'
                        # 將 YYYYMMDD 轉為 YYYYMM（取前6位）
                        filter_start = start_date[:6]
                        filter_end = end_date[:6]
                        filter_dates = (filter_start, filter_end)
                        print(f"    日期範圍: {start_date}-{end_date} → {filter_start}-{filter_end}")
                    else:
                        print(f"  ✓ 偵測到使用 formatted_time (YYYYMMDD 格式)")
                        time_field = 'formatted_time'
                        time_format = 'YYYYMMDD'
                        filter_dates = (start_date, end_date)
                    
                    # 步驟1: 尋找子查詢內的 FROM 子句
                    from_pattern = re.compile(r'FROM\s+(\w+)\.collect_record')
                    match = from_pattern.search(sql)
                    
                    if match:
                        schema_name = match.group(1)  # 取得實際的 schema 名稱
                        old_from = f"FROM {schema_name}.collect_record"
                        # 根據時間欄位類型選擇過濾條件
                        new_from = f"FROM {schema_name}.collect_record\n    WHERE to_char(to_timestamp(created_at / 1000000000), '{time_format}') BETWEEN '{filter_dates[0]}' AND '{filter_dates[1]}'"
                        
                        # 只替換第一個出現的（應該是子查詢內的）
                        sql = sql.replace(old_from, new_from, 1)
                        print(f"  ✓ 已在子查詢中添加 WHERE 條件（提前過濾）")
                        print(f"    Schema: {schema_name}")
                        print(f"    條件: {time_field} BETWEEN '{filter_dates[0]}' AND '{filter_dates[1]}'")
                        
                        # 步驟2: 移除外層的 WHERE 條件（因為已經在子查詢中過濾了）
                        old_pattern = f"WHERE cr.{time_field} BETWEEN '{filter_dates[0]}' AND '{filter_dates[1]}'"
                        
                        if old_pattern in sql:
                            # 直接移除外層的 WHERE 條件
                            sql = sql.replace(old_pattern, "")
                            print(f"  ✓ 已移除外層 WHERE 條件（避免重複過濾）")
                            message_label.config(text=f"已優化 SQL：使用 {time_field} 提前過濾", fg="green")
                            root.update()
                    else:
                        print(f"  ⚠ 警告：無法找到 FROM <schema>.collect_record 模式，不進行優化")
                        message_label.config(text="SQL 查詢執行中...", fg="blue")
                        root.update()
                else:
                    print(f"  ℹ 此 SQL 沒有子查詢結構，不進行優化（保持原樣）")
                    message_label.config(text="SQL 查詢執行中...", fg="blue")
                root.update()

        # 顯示完整 SQL 語句
        message_label.config(text=f"SQL: {sql}", fg="green")
        root.update()
        print("\n" + "=" * 100)
        print("【最終執行的 SQL 語句 - 可直接複製到 pgAdmin4 執行】:")
        print("=" * 100)
        print(sql)
        print("=" * 100)
        print("👆 上方 SQL 可直接複製到 pgAdmin4 測試")
        print("=" * 100 + "\n")

        message_label.config(text=f"連接到資料庫: {subno}", fg="blue")
        update_progress(70)
        root.update()

        conn_info = json.loads(connections[subno])
        conn_str = f"postgresql://{conn_info['user']}:{conn_info['password']}@{conn_info['host']}/{conn_info['dbname']}"
        message_label.config(text=f"連線字串: {conn_str.replace(conn_info['password'], '****')}", fg="blue")
        update_progress(80)
        root.update()

        print(f"連線到資料庫: {subno}")
        print(f"連線字串: {conn_str.replace(conn_info['password'], '****')}")

        engine = create_engine(conn_str)
        update_progress(85)
        root.update()
        
        df_result = pd.read_sql(sql, engine)
        
        update_progress(95)
        root.update()

        print(f"查詢完成，獲得 {len(df_result)} 筆資料")
        print("資料欄位:", list(df_result.columns))
        print("前5筆資料預覽:")
        print(df_result.head())

        # 清空結果區域
        for widget in result_frame.winfo_children():
            widget.destroy()

        # 建立帶有捲軸的容器
        container = tk.Frame(result_frame)
        container.pack(fill="both", expand=True)

        # 檢查是否有資料
        if len(df_result) == 0:
            message_label.config(text="查詢完成，但沒有資料 (總筆數: 0)", fg="orange")
            print("WARNING: 查詢返回0筆資料")
            # 顯示空的表格結構
            empty_df = pd.DataFrame(columns=['查詢結果', '說明'])
            empty_df.loc[0] = ['無資料', '請檢查查詢條件']
            result_table = Table(container, dataframe=empty_df, showtoolbar=True, showstatusbar=True)
        else:
            message_label.config(text=f"查詢完成 | 總筆數: {len(df_result)} 筆", fg="blue", font=("Arial", 14, "bold"))
            df_formatted, numeric_cols = format_numeric_columns(df_result)
            result_table = Table(container, dataframe=df_formatted, showtoolbar=True, showstatusbar=True)
            
            # 儲存數字欄位資訊供後續使用
            result_table.numeric_columns = numeric_cols

        result_table.show()
        
        # 調整捲軸大小
        if hasattr(result_table, 'vscrollbar'):
            result_table.vscrollbar.config(width=256)  # 垂直捲軸寬度 256 像素
        if hasattr(result_table, 'hscrollbar'):
            result_table.hscrollbar.config(width=256)  # 水平捲軸高度 256 像素
        
        # 設定表格選項和固定大小
        result_table.editable = False  # 設定為不可編輯
        
        # 禁用左右循環功能（覆寫按鍵綁定）
        result_table.unbind("<Left>")
        result_table.unbind("<Right>")
        result_table.bind("<Left>", lambda e: result_table.handle_arrow_keys(e) if result_table.currentcol > 0 else None)
        result_table.bind("<Right>", lambda e: result_table.handle_arrow_keys(e) if result_table.currentcol < len(result_table.model.df.columns) - 1 else None)
        
        result_table.redraw()

        btn_save.config(state="normal")
        btn_mail.config(state="normal")
        btn_delivery.config(state="normal")
        
        update_progress(100)
        root.update()
        
        # 2秒後清除進度
        root.after(2000, lambda: progress_label.config(text=""))

        print("SUCCESS: 表格顯示完成")
        print(f"總筆數: {len(df_result)}")
    
    except Exception as e:
        error_msg = f"執行錯誤: {e}"
        message_label.config(text=error_msg, fg="red")
        progress_label.config(text="", fg="red")
        
        # ===== 終端機顯示：錯誤資訊 =====
        print("\n" + "🔴" * 50)
        print("❌ 執行發生錯誤！")
        print("🔴" * 50)
        print(f"錯誤類型: {type(e).__name__}")
        print(f"錯誤訊息: {e}")
        print("\n詳細錯誤追蹤:")
        print("-" * 100)
        import traceback
        traceback.print_exc()
        print("-" * 100)
        print("🔴" * 50 + "\n")


def execute_d002(is_group=False):
    """D002 統計：按產品類型統計單廠或全集團的筆數和加總
    
    Args:
        is_group: True=集團查詢（所有廠別），False=單廠查詢（選擇的廠別）
    """
    global df_result, is_group_query
    is_group_query = is_group
    
    # 如果是單廠查詢，需要檢查廠別
    if not is_group:
        subno = subno_var.get()
        if not subno:
            messagebox.showerror("錯誤", "請選擇廠別！")
            return
    
    print("\n" + "=" * 100)
    if is_group:
        print("【D002 集團統計】開始執行...")
    else:
        print("【D002 單廠統計】開始執行...")
    print("=" * 100)
    
    # 獲取輸入參數
    start_month = ''
    end_month = ''
    
    if 'input1' in input_entries:
        start_month = input_entries['input1'].get()
    if 'input2' in input_entries:
        end_month = input_entries['input2'].get()
    
    print(f"獲取的輸入參數: start_month='{start_month}', end_month='{end_month}'")
    
    if not start_month or not end_month:
        messagebox.showerror("錯誤", "請輸入年月範圍！")
        print("❌ 錯誤：未輸入年月範圍")
        return
    
    # 根據查詢類型顯示訊息
    if is_group:
        message_label.config(text=f"D002 集團統計執行中... 年月: {start_month}-{end_month}", fg="blue")
        # 收集數據：count 和 sum（多廠別）
        count_data = {}  # {ptype: {factory: count}}
        sum_data = {}    # {ptype: {factory: sum}}
        # 過濾掉印尼廠（尚未建置）
        factories_to_query = [f for f in factories if f != 'I']
    else:
        subno = subno_var.get()
        print(f"選擇的廠別: {subno}")
        message_label.config(text=f"D002 單廠統計執行中... 廠別: {subno}, 年月: {start_month}-{end_month}", fg="blue")
        # 收集數據：count 和 sum（單一廠別）
        count_data = {}  # {ptype: count}
        sum_data = {}    # {ptype: sum}
        factories_to_query = [subno]
    
    root.update()
    
    # 依序查詢廠別
    total_factories = len(factories_to_query)
    
    for idx, factory in enumerate(factories_to_query, 1):
        try:
            progress_percent = int(20 + (idx / total_factories) * 60)  # 20-80%
            update_progress(progress_percent)
            
            if is_group:
                message_label.config(text=f"正在統計 [{idx}/{total_factories}] {factory}...", fg="blue")
            else:
                message_label.config(text=f"正在查詢 {factory} 數據...", fg="blue")
            root.update()
            
            print(f"\n--- 統計廠別 [{idx}/{total_factories}]: {factory} ---")
            
            # 獲取廠別設定
            selected_factory = df_factories[df_factories["subno"] == factory]
            if selected_factory.empty:
                print(f"⚠ 警告：找不到 {factory} 的設定，跳過")
                continue
            
            replace2_val = selected_factory.iloc[0]["replace2"]
            
            # 連接資料庫
            conn_info = json.loads(connections[factory])
            conn_str = f"postgresql://{conn_info['user']}:{conn_info['password']}@{conn_info['host']}/{conn_info['dbname']}"
            engine = create_engine(conn_str)
            
            # SQL: 統計筆數和加總
            sql = f"""
            SELECT 
                wo.information->>'product_type' AS ptype,
                COUNT(*) AS cnt,
                SUM(COALESCE(NULLIF(cr.detail->>'quantity','')::numeric, 0)) AS quantity_sum
            FROM (
                SELECT work_order, detail,
                       to_char(to_timestamp(created_at / 1000000000), 'YYYYMM') AS filter_time
                FROM {replace2_val}.collect_record
                WHERE to_char(to_timestamp(created_at / 1000000000), 'YYYYMM') BETWEEN '{start_month}' AND '{end_month}'
            ) AS cr
            JOIN {replace2_val}.work_order wo ON cr.work_order = wo.id
            GROUP BY wo.information->>'product_type'
            ORDER BY ptype
            """
            
            df_factory = pd.read_sql(sql, engine)
            
            if len(df_factory) > 0:
                for _, row in df_factory.iterrows():
                    ptype = row['ptype'] if pd.notna(row['ptype']) else 'NULL'
                    
                    if is_group:
                        # 集團查詢：多廠別
                        if ptype not in count_data:
                            count_data[ptype] = {}
                            sum_data[ptype] = {}
                        count_data[ptype][factory] = row['cnt']
                        sum_data[ptype][factory] = row['quantity_sum']
                    else:
                        # 單廠查詢
                        count_data[ptype] = row['cnt']
                        sum_data[ptype] = row['quantity_sum']
                
                print(f"✓ {factory} 統計成功：{len(df_factory)} 個產品類型")
            else:
                print(f"ℹ {factory} 統計完成：0 筆資料")
                
        except Exception as e:
            print(f"✗ {factory} 統計失敗：{e}")
            import traceback
            traceback.print_exc()
            if not is_group:
                messagebox.showerror("錯誤", f"查詢失敗：{e}")
                return
            continue
    
    # 轉換為 DataFrame
    print(f"\n收集到的數據：count_data 有 {len(count_data)} 個產品類型")
    
    if count_data:
        print("開始創建 DataFrame...")
        
        if is_group:
            # 集團查詢：多廠別（橫向顯示）
            df_count = pd.DataFrame(count_data).T
            df_count.fillna(0, inplace=True)
            for col in df_count.columns:
                df_count[col] = df_count[col].astype(int)
            df_count.index.name = f'{start_month}-{end_month}(筆數)'
            df_count.reset_index(inplace=True)
            # 按產品類型字母順序排序
            df_count.sort_values(by=df_count.columns[0], inplace=True)
            df_count.reset_index(drop=True, inplace=True)
            print(f"✓ 筆數表：{df_count.shape[0]} 行 x {df_count.shape[1]} 列（已按產品類型排序）")
            
            # 創建加總表
            df_sum = pd.DataFrame(sum_data).T
            df_sum.fillna(0, inplace=True)
            for col in df_sum.columns:
                df_sum[col] = df_sum[col].astype(int)
            df_sum.index.name = f'{start_month}-{end_month}(合計)'
            df_sum.reset_index(inplace=True)
            # 按產品類型字母順序排序
            df_sum.sort_values(by=df_sum.columns[0], inplace=True)
            df_sum.reset_index(drop=True, inplace=True)
            print(f"✓ 合計表：{df_sum.shape[0]} 行 x {df_sum.shape[1]} 列（已按產品類型排序）")
        else:
            # 單廠查詢：單一欄位
            subno = subno_var.get()
            df_count = pd.DataFrame(list(count_data.items()), columns=[f'{start_month}-{end_month}(筆數)', subno])
            df_count[subno] = df_count[subno].astype(int)
            # 按產品類型字母順序排序
            df_count.sort_values(by=df_count.columns[0], inplace=True)
            df_count.reset_index(drop=True, inplace=True)
            print(f"✓ 筆數表：{df_count.shape[0]} 行 x {df_count.shape[1]} 列（已按產品類型排序）")
            
            # 創建加總表（單欄，取整數）
            df_sum = pd.DataFrame(list(sum_data.items()), columns=[f'{start_month}-{end_month}(合計)', subno])
            df_sum[subno] = df_sum[subno].astype(int)
            # 按產品類型字母順序排序
            df_sum.sort_values(by=df_sum.columns[0], inplace=True)
            df_sum.reset_index(drop=True, inplace=True)
            print(f"✓ 合計表：{df_sum.shape[0]} 行 x {df_sum.shape[1]} 列（已按產品類型排序）")
        
        # 使用 Notebook 顯示兩個表格
        for widget in result_frame.winfo_children():
            widget.destroy()
        
        # 設定 Notebook 樣式（原始大小字體、黃色 active、灰色 inactive）
        style = ttk.Style()
        style.theme_use('default')
        style.configure('D002.TNotebook.Tab', 
                       padding=[10, 5],
                       background='#D3D3D3')  # 灰色底
        style.map('D002.TNotebook.Tab',
                 background=[('selected', '#FFFF00')])  # active 黃色底
        
        notebook = ttk.Notebook(result_frame, style='D002.TNotebook')
        notebook.pack(fill="both", expand=True)
        
        # 工作表1：筆數（整數 + 千位數符號）
        tab_count = tk.Frame(notebook)
        notebook.add(tab_count, text="筆數")
        # 轉換數字列為帶千位數符號的字串
        df_count_display = df_count.copy()
        for col in df_count_display.columns:
            if col != df_count_display.columns[0]:  # 跳過第一欄（產品類型）
                if pd.api.types.is_numeric_dtype(df_count_display[col]):
                    df_count_display[col] = df_count_display[col].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "")
        table_count = Table(tab_count, dataframe=df_count_display, showtoolbar=True, showstatusbar=True)
        table_count.show()
        print("✓ 筆數表已顯示（整數 + 千位數符號）")
        
        # 工作表2：合計（整數 + 千位數符號）
        tab_sum = tk.Frame(notebook)
        notebook.add(tab_sum, text="合計")
        # 轉換數字列為帶千位數符號的字串
        df_sum_display = df_sum.copy()
        for col in df_sum_display.columns:
            if col != df_sum_display.columns[0]:  # 跳過第一欄（產品類型）
                if pd.api.types.is_numeric_dtype(df_sum_display[col]):
                    df_sum_display[col] = df_sum_display[col].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "")
        table_sum = Table(tab_sum, dataframe=df_sum_display, showtoolbar=True, showstatusbar=True)
        table_sum.show()
        print("✓ 合計表已顯示（整數 + 千位數符號）")
        
        # 儲存結果供存檔和郵件使用
        df_result = {'筆數': df_count, '合計': df_sum}
        
        # 啟用按鈕
        btn_save.config(state="normal")
        btn_mail.config(state="normal")
        btn_delivery.config(state="normal")
        
        update_progress(100)
        if is_group:
            message_label.config(
                text=f"✓ D002 集團統計完成！共 {len(df_count)} 個產品類型",
                fg="green"
            )
            print("\n" + "=" * 100)
            print(f"【D002 集團統計】完成！")
            print(f"產品類型數量：{len(df_count)}")
            print("=" * 100)
        else:
            message_label.config(
                text=f"✓ D002 單廠統計完成！共 {len(df_count)} 個產品類型",
                fg="green"
            )
            print("\n" + "=" * 100)
            print(f"【D002 單廠統計】完成！")
            print(f"產品類型數量：{len(df_count)}")
            print("=" * 100)
        
        root.after(2000, lambda: progress_label.config(text=""))
    else:
        message_label.config(text="D002 統計完成，但沒有任何資料", fg="orange")
        progress_label.config(text="")


def execute_group_sql():
    """集團查詢：依序查詢所有廠別並合併結果"""
    global df_result
    
    print("\n" + "=" * 100)
    print("【集團查詢】開始執行...")
    print("=" * 100)
    
    selected_code = code_var.get()
    code_key = code_map.get(selected_code)
    
    if not code_key:
        messagebox.showerror("錯誤", "請選擇查詢代碼！")
        return
    
    # 特殊處理：D002 集團查詢
    if code_key == 'D002':
        execute_d002(is_group=True)
        return
    
    # 檢查輸入欄位
    missing_inputs = [key for key, entry in input_entries.items() if not entry.get()]
    if missing_inputs:
        message_label.config(text=f"以下欄位尚未輸入: {', '.join(missing_inputs)}", fg="red")
        return
    
    # 初始化
    all_results = []
    total_factories = len(factories)
    success_count = 0
    
    message_label.config(text=f"集團查詢開始... 共 {total_factories} 個廠別", fg="blue")
    root.update()
    
    # 依序查詢每個廠別
    for idx, factory in enumerate(factories, 1):
        try:
            progress_percent = int((idx / total_factories) * 100)
            update_progress(progress_percent)
            message_label.config(text=f"正在查詢 [{idx}/{total_factories}] {factory}...", fg="blue")
            root.update()
            
            print(f"\n--- 查詢廠別 [{idx}/{total_factories}]: {factory} ---")
            
            # 取得廠別設定
            selected_factory = df_factories[df_factories["subno"] == factory]
            if selected_factory.empty:
                print(f"⚠ 警告：找不到 {factory} 的設定，跳過")
                continue
            
            replace1_val = selected_factory.iloc[0]["replace1"]
            replace2_val = selected_factory.iloc[0]["replace2"]
            
            # 準備 SQL
            sql = sql_templates[code_key]
            
            # 替換 subno 和 factory（陣列方式多次替換）
            if pd.notna(replace1_val):
                print(f"  替換 subno/factory (陣列方式):")
                原始陣列 = [("'7' as subno", "'7' AS subno"), ("'L' as factory", "'L' AS factory")]
                
                try:
                    新陣列 = [item.strip() for item in replace1_val.split(',')]
                    
                    if len(新陣列) == len(原始陣列):
                        for i, (舊值組, 新值) in enumerate(zip(原始陣列, 新陣列), 1):
                            舊值_小寫, 舊值_大寫 = 舊值組
                            print(f"    替換 {i}: '{舊值_小寫}' 或 '{舊值_大寫}' → '{新值}'")
                            sql = sql.replace(舊值_小寫, 新值)
                            sql = sql.replace(舊值_大寫, 新值)
                    else:
                        print(f"    ⚠ 陣列長度不一致，使用原始方式替換")
                        sql = sql.replace("'7' as subno,'L' as factory", replace1_val)
                except Exception as e:
                    print(f"    ⚠ 拆分失敗，使用原始方式替換 - {e}")
                    sql = sql.replace("'7' as subno,'L' as factory", replace1_val)
            
            # 替換 schema
            if pd.notna(replace2_val):
                print(f"  替換 schema: kt2mes → {replace2_val}")
                sql = sql.replace("kt2mes", replace2_val)
            
            # 替換輸入參數
            for key, entry in input_entries.items():
                value = entry.get()
                if value:
                    sql = sql.replace(key, f"'{value}'")
            
            # SQL 優化（日期查詢）
            if is_date_parameter('input1') and 'input1' in input_entries and 'input2' in input_entries:
                start_date = input_entries['input1'].get()
                end_date = input_entries['input2'].get()
                
                if start_date and end_date and len(start_date) == 8 and len(end_date) == 8:
                    has_subquery = re.search(r'FROM\s*\(\s*SELECT.*?\)\s*AS\s+cr', sql, re.DOTALL | re.IGNORECASE)
                    
                    if has_subquery:
                        # 檢測 SQL 使用哪種時間欄位
                        uses_filter_time = 'cr.filter_time' in sql or 'filter_time AS' in sql
                        
                        if uses_filter_time:
                            print(f"  ✓ 使用 filter_time (YYYYMM)")
                            time_field = 'filter_time'
                            time_format = 'YYYYMM'
                            filter_start = start_date[:6]
                            filter_end = end_date[:6]
                            filter_dates = (filter_start, filter_end)
                        else:
                            print(f"  ✓ 使用 formatted_time (YYYYMMDD)")
                            time_field = 'formatted_time'
                            time_format = 'YYYYMMDD'
                            filter_dates = (start_date, end_date)
                        
                        from_pattern = re.compile(r'FROM\s+(\w+)\.collect_record')
                        match = from_pattern.search(sql)
                        
                        if match:
                            schema_name = match.group(1)
                            old_from = f"FROM {schema_name}.collect_record"
                            new_from = f"FROM {schema_name}.collect_record\n    WHERE to_char(to_timestamp(created_at / 1000000000), '{time_format}') BETWEEN '{filter_dates[0]}' AND '{filter_dates[1]}'"
                            sql = sql.replace(old_from, new_from, 1)
                            
                            old_pattern = f"WHERE cr.{time_field} BETWEEN '{filter_dates[0]}' AND '{filter_dates[1]}'"
                            if old_pattern in sql:
                                sql = sql.replace(old_pattern, "")
                                print(f"  ✓ 已優化（{time_field}: {filter_dates[0]}-{filter_dates[1]}）")
            
            # 顯示最終 SQL
            print("\n" + "-" * 80)
            print(f"【{factory} 的 SQL 語句 - 可直接複製到 pgAdmin4 執行】:")
            print("-" * 80)
            print(sql)
            print("-" * 80)
            
            # 連接資料庫
            conn_info = json.loads(connections[factory])
            conn_str = f"postgresql://{conn_info['user']}:{conn_info['password']}@{conn_info['host']}/{conn_info['dbname']}"
            print(f"  連線: {conn_str.replace(conn_info['password'], '****')}")
            engine = create_engine(conn_str)
            
            # 執行查詢
            df_factory = pd.read_sql(sql, engine)
            
            if len(df_factory) > 0:
                all_results.append(df_factory)
                success_count += 1
                print(f"✓ {factory} 查詢成功：{len(df_factory)} 筆資料")
            else:
                print(f"ℹ {factory} 查詢完成：0 筆資料")
                
        except Exception as e:
            print(f"✗ {factory} 查詢失敗：{e}")
            continue
    
    # 合併結果
    if all_results:
        df_result = pd.concat(all_results, ignore_index=True)
        
        # 清空結果區域
        for widget in result_frame.winfo_children():
            widget.destroy()
        
        # 建立帶有捲軸的容器
        container = tk.Frame(result_frame)
        container.pack(fill="both", expand=True)
        
        # 格式化並顯示
        df_formatted, numeric_cols = format_numeric_columns(df_result)
        result_table = Table(container, dataframe=df_formatted, showtoolbar=True, showstatusbar=True)
        result_table.numeric_columns = numeric_cols
        result_table.show()
        
        # 調整捲軸
        if hasattr(result_table, 'vscrollbar'):
            result_table.vscrollbar.config(width=256)
        if hasattr(result_table, 'hscrollbar'):
            result_table.hscrollbar.config(width=256)
        
        result_table.editable = False
        result_table.unbind("<Left>")
        result_table.unbind("<Right>")
        result_table.bind("<Left>", lambda e: result_table.handle_arrow_keys(e) if result_table.currentcol > 0 else None)
        result_table.bind("<Right>", lambda e: result_table.handle_arrow_keys(e) if result_table.currentcol < len(result_table.model.df.columns) - 1 else None)
        result_table.redraw()
        
        # 啟用按鈕
        btn_save.config(state="normal")
        btn_mail.config(state="normal")
        btn_delivery.config(state="normal")
        
        update_progress(100)
        message_label.config(
            text=f"✓ 集團查詢完成！成功查詢 {success_count}/{total_factories} 個廠別 | 總筆數：{len(df_result)} 筆",
            fg="green"
        )
        
        print("\n" + "=" * 100)
        print(f"【集團查詢】完成！")
        print(f"成功查詢：{success_count}/{total_factories} 個廠別")
        print(f"總資料筆數：{len(df_result)}")
        print("=" * 100)
        
        root.after(2000, lambda: progress_label.config(text=""))
    else:
        message_label.config(text="集團查詢完成，但沒有任何資料", fg="orange")
        progress_label.config(text="")


result_frame = tk.Frame(root, borderwidth=1, relief="solid")
result_frame.place(x=0, y=150, relwidth=1, relheight=1, height=-220)

# 6. button_frame: 按鈕列的元件與函式
# --------------------------------------------------

def save_excel():
    """將查詢結果儲存為 Excel 檔案。"""
    global df_result
    if df_result is not None:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")],
                                                 title="儲存 Excel 檔案")
        if file_path:
            # 檢查是否為 D002 的多工作表格式
            if isinstance(df_result, dict):
                # D002 多工作表格式
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    for sheet_name, df in df_result.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                message_label.config(text=f"Excel 檔案（多工作表）已儲存至：{file_path}", fg="blue")
            else:
                # 一般單工作表格式
                df_result.to_excel(file_path, index=False)
                message_label.config(text=f"Excel 檔案已儲存至：{file_path}", fg="blue")
            save_excel.last_saved_path = file_path

def send_query_result_email():
    """自動將查詢結果儲存為 Excel 並寄送郵件。"""
    global df_result
    
    if df_result is None or len(df_result) == 0:
        messagebox.showwarning("無資料", "目前沒有查詢結果可以寄送！\n請先執行查詢。")
        return
    
    try:
        # 1. 自動產生暫存檔案名稱
        import tempfile
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        廠別 = subno_var.get().split('.')[0]  # 取得廠別代碼，如 'T'
        查詢代碼 = code_var.get().split()[0] if code_var.get() else "查詢"  # 取得查詢代碼
        
        temp_dir = tempfile.gettempdir()
        file_name = f"ERP查詢結果_{廠別}_{查詢代碼}_{timestamp}.xlsx"
        file_path = os.path.join(temp_dir, file_name)
        
        # 2. 儲存查詢結果為 Excel
        if isinstance(df_result, dict):
            # D002 多工作表格式
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in df_result.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # 一般單工作表格式
            df_result.to_excel(file_path, index=False)
        message_label.config(text=f"正在準備郵件... 檔案：{file_name}", fg="blue")
        root.update()
        
        # 3. 準備郵件
        recipient = f"{username}@kenda.com.tw"
        bcc = "oak@kenda.com.tw"
        
        msg = EmailMessage()
        msg["Subject"] = f"ERP查詢結果 - {廠別}廠 {查詢代碼} ({timestamp})"
        msg["From"] = "erp.system@kenda.com.tw"
        msg["To"] = recipient
        msg["Bcc"] = bcc  # 密件副本
        
        # 郵件內容 - 收集輸入欄位資訊
        查詢條件 = []
        code_key = code_map.get(code_var.get())
        
        # 取得輸入欄位的標籤和值
        if code_key and code_key in input_labels and code_key in input_keys:
            try:
                labels = ast.literal_eval(input_labels[code_key])
                keys = ast.literal_eval(input_keys[code_key])
                
                for label_text, key in zip(labels, keys):
                    if key in input_entries:
                        value = input_entries[key].get()
                        if value:  # 只顯示有輸入值的欄位
                            查詢條件.append(f"  - {label_text}: {value}")
            except:
                pass
        
        # 組合查詢條件字串
        查詢條件文字 = "\n".join(查詢條件) if 查詢條件 else "  (無輸入條件)"
        
        # 判斷是集團查詢還是單一廠別查詢
        code_key = code_map.get(code_var.get())
        if code_key == 'D002' or is_group_query:
            廠別顯示 = "全集團各廠"
        elif isinstance(df_result, dict):
            # D002 格式
            廠別顯示 = "全集團各廠"
        elif hasattr(df_result, 'columns') and 'subno' in df_result.columns and 'factory' in df_result.columns:
            # 檢查 df_result 中是否有多個不同的廠別
            unique_factories = df_result[['subno', 'factory']].drop_duplicates()
            if len(unique_factories) > 1:
                廠別顯示 = "全集團各廠"
            else:
                廠別顯示 = subno_var.get()
        else:
            廠別顯示 = subno_var.get()
        
        # 計算資料筆數
        if isinstance(df_result, dict):
            # D002 多工作表格式
            total_rows = sum(len(df) for df in df_result.values())
            資料筆數 = f"{total_rows} 筆（共 {len(df_result)} 個工作表）"
        else:
            # 一般單工作表格式
            資料筆數 = f"{len(df_result)} 筆"
        
        郵件內容 = f"""您好，

這是您在 {timestamp[:4]}/{timestamp[4:6]}/{timestamp[6:8]} {timestamp[9:11]}:{timestamp[11:13]}:{timestamp[13:15]} 執行的 ERP 查詢結果。

查詢資訊：
- 廠別：{廠別顯示}
- 查詢代碼：{code_var.get()}
- 查詢條件：
{查詢條件文字}
- 資料筆數：{資料筆數}

詳細資料請參閱附件 Excel 檔案。

---
此郵件由 ERP 查詢系統自動發送
"""
        msg.set_content(郵件內容)
        
        # 4. 附加 Excel 檔案
        with open(file_path, "rb") as f:
            file_data = f.read()
            msg.add_attachment(
                file_data,
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=file_name
            )
        
        # 5. 發送郵件
        message_label.config(text=f"正在發送郵件至 {recipient}...", fg="blue")
        root.update()
        
        with smtplib.SMTP("192.1.1.241", 25) as server:
            # server.login("tiptop", "tiptop")  # 如果需要認證
            server.send_message(msg)
        
        # 6. 完成
        message_label.config(
            text=f"✓ 郵件已成功寄出！\n收件人：{recipient} | 密件副本：{bcc} | 附件：{file_name}",
            fg="green"
        )
        messagebox.showinfo(
            "郵件已寄出",
            f"查詢結果已成功寄送！\n\n收件人：{recipient}\n密件副本：{bcc}\n附件：{file_name}\n資料筆數：{len(df_result)} 筆"
        )
        
        # 7. 清理暫存檔案（可選）
        try:
            os.remove(file_path)
        except:
            pass
            
    except Exception as e:
        error_msg = f"郵件發送失敗：{e}"
        message_label.config(text=error_msg, fg="red")
        messagebox.showerror("郵件發送失敗", f"發生錯誤：\n{e}")
        print(f"ERROR: {error_msg}")
        import traceback
        traceback.print_exc()

def send_email():
    """發送查詢結果 Excel 檔案至使用者信箱。"""
    try:
        recipient = f"{username}@kenda.com.tw"
        msg = EmailMessage()
        msg["Subject"] = "ERP查詢結果"
        msg["From"] = "noreply@kenda.com.tw"
        msg["To"] = recipient
        msg.set_content("附件為查詢結果")
        file_path = getattr(save_excel, "last_saved_path", None)
        if not file_path:
            message_label.config(text="尚未儲存 Excel 檔案，無法寄送", fg="red")
            return
        with open(file_path, "rb") as f:
            msg.add_attachment(f.read(), maintype="application", subtype="octet-stream", filename=os.path.basename(file_path))
        with smtplib.SMTP("192.1.1.241", 25) as server:
            server.send_message(msg)
        message_label.config(text=f"郵件已寄出至 {recipient}", fg="blue")
    except Exception as e:
        message_label.config(text=f"錯誤: {e}", fg="red")

def yaml_to_excel():
    """讓使用者選擇 YAML 檔案並將其轉換為 Excel。"""
    global df_result
    yaml_file = filedialog.askopenfilename(
        title="選擇要轉換的 YAML 文件",
        filetypes=[("YAML files", "*.yaml")]
    )
    if not yaml_file:
        return  # 使用者取消了選擇
    
    # 建立與 YAML 檔案同名的 Excel 檔案路徑
    excel_file = os.path.splitext(yaml_file)[0] + '.xlsx'

    try:
        with open(yaml_file, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f)
            workbook = openpyxl.Workbook()
            sheet = workbook.active
        
        # 遞迴函數來寫入數據
        def write_data(data, row_offset, col_offset, parent_key=''):
            if isinstance(data, dict):
                for key, value in data.items():
                    cell_key = f"{parent_key}.{key}" if parent_key else key
                    sheet.cell(row=row_offset, column=col_offset).value = cell_key
                    if isinstance(value, (dict, list)):
                        row_offset = write_data(value, row_offset + 1, col_offset, cell_key)
                    else:
                        sheet.cell(row=row_offset, column=col_offset + 1).value = value
                        row_offset += 1
                return row_offset
            elif isinstance(data, list):
                for item in data:
                    row_offset = write_data(item, row_offset, col_offset + 1, parent_key)
                return row_offset
            else:
                sheet.cell(row=row_offset, column=col_offset).value = data
                return row_offset + 1

        write_data(data, 1, 1)
        workbook.save(excel_file)
        
        # 讀取剛生成的 Excel 文件並顯示在表格中
        df_result = pd.read_excel(excel_file, engine='openpyxl')
        for widget in result_frame.winfo_children():
            widget.destroy()
        
        # 建立帶有捲軸的容器
        container = tk.Frame(result_frame)
        container.pack(fill="both", expand=True)
        
        df_formatted, numeric_cols = format_numeric_columns(df_result)
        result_table = Table(container, dataframe=df_formatted, showtoolbar=True, showstatusbar=True)
        result_table.numeric_columns = numeric_cols
        result_table.show()
        
        # 調整捲軸大小
        if hasattr(result_table, 'vscrollbar'):
            result_table.vscrollbar.config(width=256)  # 垂直捲軸寬度 256 像素
        if hasattr(result_table, 'hscrollbar'):
            result_table.hscrollbar.config(width=256)  # 水平捲軸高度 256 像素
        
        # 設定表格選項和固定大小
        result_table.editable = False  # 設定為不可編輯
        
        # 禁用左右循環功能（覆寫按鍵綁定）
        result_table.unbind("<Left>")
        result_table.unbind("<Right>")
        result_table.bind("<Left>", lambda e: result_table.handle_arrow_keys(e) if result_table.currentcol > 0 else None)
        result_table.bind("<Right>", lambda e: result_table.handle_arrow_keys(e) if result_table.currentcol < len(result_table.model.df.columns) - 1 else None)
        
        result_table.redraw()

        # 更新訊息標籤和按鈕狀態
        message_label.config(text=f"已經產生出, '{excel_file}', '{os.path.basename(excel_file)}'", fg="blue")
        btn_save.config(state="normal")
        btn_delivery.config(state="normal")

    except FileNotFoundError:
        message_label.config(text=f"錯誤: 找不到 '{yaml_file}' 文件", fg="red")
    except yaml.YAMLError as e:
        message_label.config(text=f"YAML 錯誤: {e}", fg="red")
    except Exception as e:
        message_label.config(text=f"發生錯誤: {e}", fg="red")

def excel_to_yaml():
    """讓使用者選擇 Excel 檔案並將其轉換為 YAML。"""
    global df_result
    excel_file = filedialog.askopenfilename(
        title="選擇要轉換的 Excel 文件",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not excel_file:
        return  # 使用者取消了選擇

    # 建立與 Excel 檔案同名的 YAML 檔案路徑
    yaml_file = os.path.splitext(excel_file)[0] + '.yaml'

    try:
        # 讀取 Excel 檔案
        df_result = pd.read_excel(excel_file, engine='openpyxl')
        
        # 將 DataFrame 轉換為字典列表，以方便轉換為 YAML
        # 處理 NaN 值，將其轉換為 None 或空字串
        data = df_result.replace({pd.NA: None}).to_dict(orient='records')
        
        # 寫入 YAML 檔案
        with open(yaml_file, 'w', encoding='utf-8') as f:
            yaml.dump(data, f, allow_unicode=True)

        # 更新訊息標籤
        message_label.config(text=f"已成功將 '{os.path.basename(excel_file)}' 轉換為 '{os.path.basename(yaml_file)}'", fg="blue")
        
        # 將結果顯示在表格中 (可選)
        for widget in result_frame.winfo_children():
            widget.destroy()
        
        # 建立帶有捲軸的容器
        container = tk.Frame(result_frame)
        container.pack(fill="both", expand=True)
        
        # 使用PandasTable顯示結果
        df_formatted, numeric_cols = format_numeric_columns(df_result)
        result_table = Table(container, dataframe=df_formatted, showtoolbar=True, showstatusbar=True)
        result_table.numeric_columns = numeric_cols
        result_table.show()
        
        # 調整捲軸大小
        if hasattr(result_table, 'vscrollbar'):
            result_table.vscrollbar.config(width=256)  # 垂直捲軸寬度 256 像素
        if hasattr(result_table, 'hscrollbar'):
            result_table.hscrollbar.config(width=256)  # 水平捲軸高度 256 像素
        
        # 設定表格選項和固定大小
        result_table.editable = False  # 設定為不可編輯
        
        # 禁用左右循環功能（覆寫按鍵綁定）
        result_table.unbind("<Left>")
        result_table.unbind("<Right>")
        result_table.bind("<Left>", lambda e: result_table.handle_arrow_keys(e) if result_table.currentcol > 0 else None)
        result_table.bind("<Right>", lambda e: result_table.handle_arrow_keys(e) if result_table.currentcol < len(result_table.model.df.columns) - 1 else None)
        
        result_table.redraw()
        
        # 更新按鈕狀態
        btn_save.config(state="normal")
        btn_delivery.config(state="normal")

    except FileNotFoundError:
        message_label.config(text=f"錯誤: 找不到 '{excel_file}' 文件", fg="red")
    except Exception as e:
        message_label.config(text=f"發生錯誤: {e}", fg="red")

def quit_app():
    """退出程式。"""
    root.quit()
    root.destroy()

button_frame = tk.Frame(root, borderwidth=1, relief="flat")
button_frame.place(x=0, y=90, width=screen_width, height=50)

# 按鈕間距統一為 100 像素
btn_query = tk.Button(button_frame, text="查詢", command=execute_sql, font=(12), width=8, height=2)
btn_query.place(x=10, y=5)

btn_group_query = tk.Button(button_frame, text="集團查詢", command=execute_group_sql, font=(12), width=8, height=2)
btn_group_query.place(x=110, y=5)

btn_mail = tk.Button(button_frame, text="寄送Mail", command=send_query_result_email, font=(12), width=8, height=2, state="disabled")
btn_mail.place(x=210, y=5)

btn_save = tk.Button(button_frame, text="Excel存檔", command=save_excel, font=(12), width=8, height=2, state="disabled")
btn_save.place(x=310, y=5)

btn_delivery = tk.Button(button_frame, text="發送E-mail", command=send_email, font=(12), width=8, height=2, state="disabled")
# btn_delivery.place(x=310, y=5)  # 暫時隱藏（舊的郵件功能）

btn_sql = tk.Button(button_frame, text="SQL語法", command=show_sql, font=(12), width=8, height=2)
# btn_sql.place(x=410, y=5)  # 暫時隱藏

btn_condi = tk.Button(button_frame, text="條件說明", command=show_note, font=(12), width=8, height=2)
# btn_condi.place(x=510, y=5)  # 暫時隱藏

# btn_yaml = tk.Button(button_frame, text="Yaml轉Excel", command=yaml_to_excel, font=(12), width=10, height=2)
# btn_yaml.place(x=410, y=5)

# btn_excel = tk.Button(button_frame, text="Excel轉Yaml", command=excel_to_yaml, font=(12), width=10, height=2)
# btn_excel.place(x=520, y=5)

btn_quit = tk.Button(button_frame, text="退出", command=quit_app, font=(12), width=8, height=2)
btn_quit.place(x=1430, y=5)

# 7. 結束程式必須的公用語法與函式
# --------------------------------------------------

code_combo.bind("<<ComboboxSelected>>", update_inputs)
update_inputs()

root.mainloop()